import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import matplotlib.pyplot as plt

# Function to check if a file exists
def file_exists(filename):
    return os.path.isfile(filename)

# Function to collect data and store it in an Excel file
def collect_and_store_data(filename):
    try:
        # Create a new Excel workbook and select the active sheet
        if not file_exists(filename):
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Date", "Type of Material", "Quantity", "Price"]
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                cell = sheet[f"{column_letter}1"]
                cell.value = header
                cell.font = Font(bold=True)
        else:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

        # Collect data and store it in the Excel sheet....................................................
        while True:
            msg = input("To add data, enter 'a', to finish, enter 'f': ")

            if msg.lower() == 'a':
                material = input("Enter the type of material: ")
                date = datetime.datetime.now().strftime("%Y-%m-%d")
                quantity = int(input("Enter the quantity: "))
                price = float(input("Enter the price: "))

                # Add data to the Excel sheet...................
                row_data = [date, material, quantity, price]
                sheet.append(row_data)
            elif msg.lower() == 'f':
                break

        # Save the Excel file..............................................
        workbook.save(filename)
        print(f"Data saved to {filename}")


        # .........................................................................................
        # Create a simple bar chart from the collected data
        materials = []
        quantities = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4, values_only=True):
            materials.append(row[0])
            quantities.append(row[2])

        plt.figure(figsize=(10, 6))
        plt.bar(materials, quantities, color='skyblue')
        plt.xlabel('material')
        plt.ylabel('Quantity')
        plt.title('chart')
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.show()

    except Exception as e:
        print(f"An error occurred: {e}")


#........................................................................




if __name__ == "__main__":
    excel_filename = "Ganesh.xlsx"  # Change this to your desired filename
    collect_and_store_data(excel_filename)
